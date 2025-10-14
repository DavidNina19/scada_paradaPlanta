from redis_connect import MQTTWithRedis
from api_alex_connect import Api
import pandas as pd
import time
from datetime import datetime

def datos_alex_proceso():
    url = 'http://192.168.252.6/serviciowebaccesonet/api/authentication/logintercero'
    filas = []
    for i in range(1, 3):
        urlMaquina = f'http://192.168.252.6/serviciowebtercerosnet/api/MaquinaOrdenParte/{i}'
        api = Api(url, urlMaquina)
        data = api.get_all_data()

        if not data:
            continue

        import pandas as pd
        # Normalizar distintos tipos de retorno
        if isinstance(data, pd.DataFrame):
            filas.extend(data.to_dict('records'))
        elif isinstance(data, dict):
            filas.append(data)
        else:  # lista de dicts u otros iterables
            try:
                filas.extend(list(data))
            except Exception:
                filas.append(data)

    if not filas:
        return None

    import pandas as pd
    df = pd.DataFrame(filas)
    return df

def datos_paradas_scada():
    paradaScada = MQTTWithRedis()
    df = paradaScada.paradas_produccion()
    return df

def extraer_datos_alex():
    df = datos_alex_proceso()
    columnas_objetivo = ['codMaquina','idNumOrd', 'inicioParada', 'desParada', 'operario']

    if df is None or df.empty:
        return pd.DataFrame(columns=columnas_objetivo)

    cols_map = {}
    cols_lower = {c.lower(): c for c in df.columns}
    for target in columnas_objetivo:
        t_lower = target.lower()
        if t_lower in cols_lower:
            cols_map[target] = cols_lower[t_lower]
            continue
        candidates = [c for c in df.columns if t_lower in c.lower()]
        cols_map[target] = candidates[0] if candidates else None

    result = {}
    for target, src in cols_map.items():
        if src:
            result[target] = df[src]
        else:
            result[target] = pd.Series([pd.NA] * len(df), index=df.index)

    df_res = pd.DataFrame(result).reset_index(drop=True)

    if 'inicioParada' in df_res.columns:
        df_res['inicioParada'] = pd.to_datetime(df_res['inicioParada'], errors='coerce')

    df_res = df_res.dropna(subset=['inicioParada']).reset_index(drop=True)

    return df_res

def match_alex_scada_paradas():
    """
    Integra extraer_datos_alex() con datos_paradas_scada().
    Por cada máquina (agrupado por código de máquina normalizado) busca filas de Alex cuyo
    inicioParada cumpla cualquiera de las condiciones respecto a un periodo SCADA:
      - esté dentro del rango [inicioParada_scada, finParada_scada], o
      - esté a <= 10 minutos del inicioParada_scada.
    Devuelve un DataFrame con las coincidencias (columnas de Alex + columnas scada_inicio/scada_fin).
    """
    import re
    from datetime import timedelta

    df_alex = extraer_datos_alex()
    df_scada = datos_paradas_scada()

    # columnas objetivo de salida
    out_cols = [
        'codMaquina', 'idNumOrd', 'inicioParada_alex', 'desParada', 'operario',
        'inicioParada_scada', 'finParada_scada'
    ]

    # Validaciones iniciales
    if df_alex is None or df_alex.empty or df_scada is None or df_scada.empty:
        return pd.DataFrame(columns=out_cols)

    # Normalizador simple de nombres de máquina (igual lógica que limpiar_codmaq)
    patron_busqueda = re.compile(r"MicroWin\.PLC\d+", flags=re.IGNORECASE)
    def normalize_machine_name(val):
        if pd.isna(val):
            return ''
        s = str(val)
        s = patron_busqueda.sub("", s)
        s = s.replace(".CTO", "")
        return s.strip().lower()

    # Detectar columna de máquina en SCADA (intentar varios candidatos)
    scada_machine_col = None
    for cand in ['codmaq', 'codmaquina', 'codMaquina', 'codmaq.']:
        if cand in df_scada.columns:
            scada_machine_col = cand
            break
    if scada_machine_col is None:
        # fallback: buscar por nombre que contenga 'codm' o 'maquina'
        for c in df_scada.columns:
            low = c.lower()
            if 'codm' in low or 'maquina' in low or 'codmaq' in low:
                scada_machine_col = c
                break
    if scada_machine_col is None:
        # si no hay columna de máquina, no se puede emparejar
        return pd.DataFrame(columns=out_cols)

    # Asegurar columnas datetime en SCADA
    df_scada = df_scada.copy()
    df_scada['inicioParada_scada'] = pd.to_datetime(df_scada.get('inicioParada', df_scada.get('inicioparada', df_scada.get('inicioParada'))), errors='coerce')
    df_scada['finParada_scada'] = pd.to_datetime(df_scada.get('finParada', df_scada.get('finparada', df_scada.get('finParada'))), errors='coerce')

    # Normalizar nombres de máquina en ambos dataframes
    # Para Alex, buscar columna codMaquina o similar
    alex_machine_col = None
    for c in df_alex.columns:
        if c.lower().startswith('codm') or 'codmaq' in c.lower() or 'maquina' in c.lower():
            alex_machine_col = c
            break
    if alex_machine_col is None:
        # si Alex no tiene columna de máquina, no se puede emparejar
        return pd.DataFrame(columns=out_cols)

    df_alex = df_alex.copy()
    df_alex['machine_norm'] = df_alex[alex_machine_col].apply(normalize_machine_name)
    df_scada['machine_norm'] = df_scada[scada_machine_col].apply(normalize_machine_name)

    # Asegurar inicioParada de Alex como datetime
    df_alex['inicioParada_alex'] = pd.to_datetime(df_alex.get('inicioParada', df_alex.get('inicioparada', '')), errors='coerce')

    ten_min = pd.Timedelta(minutes=10)

    matches = []
    # iterar por máquina en la intersección
    maquinas = set(df_alex['machine_norm'].unique()) & set(df_scada['machine_norm'].unique())
    if not maquinas:
        return pd.DataFrame(columns=out_cols)

    for maq in maquinas:
        sub_alex = df_alex[df_alex['machine_norm'] == maq].copy()
        sub_scada = df_scada[df_scada['machine_norm'] == maq].copy()

        # ignorar si no hay datos temporales válidos
        sub_alex = sub_alex.dropna(subset=['inicioParada_alex'])
        if sub_alex.empty or sub_scada.empty:
            continue

        # para cada registro de Alex buscar si coincide con algún registro SCADA de la misma máquina
        for a_row in sub_alex.itertuples(index=False):
            a_inicio = a_row.inicioParada_alex
            matched = False
            for s_row in sub_scada.itertuples(index=False):
                s_inicio = getattr(s_row, 'inicioParada_scada', pd.NaT)
                s_fin = getattr(s_row, 'finParada_scada', pd.NaT)

                within_range = pd.notna(s_inicio) and pd.notna(s_fin) and (s_inicio <= a_inicio <= s_fin)
                near_start = pd.notna(s_inicio) and (abs(a_inicio - s_inicio) <= ten_min)

                if within_range or near_start:
                    matches.append({
                        'codMaquina': getattr(a_row, alex_machine_col) if alex_machine_col in df_alex.columns else getattr(a_row, 'machine_norm'),
                        'idNumOrd': a_row.idNumOrd if 'idNumOrd' in df_alex.columns else a_row.idNumOrden if 'idNumOrden' in df_alex.columns else pd.NA,
                        'inicioParada_alex': a_inicio,
                        'desParada': getattr(a_row, 'desParada', pd.NA),
                        'operario': getattr(a_row, 'operario', pd.NA),
                        'inicioParada_scada': s_inicio,
                        'finParada_scada': s_fin
                    })
                    matched = True
                    # no hacer break si quieres permitir que una fila Alex empareje múltiples periodos SCADA;
                    # si prefieres solo el primer match, descomenta el siguiente line:
                    # break
            # si no matched no se añade

    if not matches:
        return pd.DataFrame(columns=out_cols)

    df_matches = pd.DataFrame(matches)
    # normalizar tipos datetime y ordenar
    df_matches['inicioParada_alex'] = pd.to_datetime(df_matches['inicioParada_alex'], errors='coerce')
    df_matches['inicioParada_scada'] = pd.to_datetime(df_matches['inicioParada_scada'], errors='coerce')
    df_matches['finParada_scada'] = pd.to_datetime(df_matches['finParada_scada'], errors='coerce')

    return df_matches.reset_index(drop=True)

# ...existing code...
import os
from openpyxl import Workbook, load_workbook

def monitor_match_and_log_excel(interval_seconds: int = 5, output_file: str = "registro.xlsx", sheet_name: str = "registro"):
    seen = set()
    try:
        while True:
            df_matches = match_alex_scada_paradas()
            if df_matches is None:
                df_matches = pd.DataFrame()

            rows = [tuple(r) for r in df_matches.itertuples(index=False, name=None)]
            new_rows = [r for r in rows if r not in seen]

            if new_rows:
                for r in new_rows:
                    seen.add(r)

                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # preparar libro/hoja
                if not os.path.exists(output_file):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    wb = load_workbook(output_file)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)

                # Escribir marca de tiempo
                ws.append([f"--- Actualización: {ts} ---"])
                # Escribir header
                if not df_matches.empty:
                    ws.append(list(df_matches.columns))
                    # Escribir filas
                    for r in df_matches.itertuples(index=False, name=None):
                        # convertir valores tipo Timestamp a str para evitar problemas
                        row_vals = [ (v.strftime("%Y-%m-%d %H:%M:%S") if hasattr(v, "strftime") else v) for v in r ]
                        ws.append(row_vals)
                else:
                    ws.append(["No hay coincidencias en este ciclo."])

                wb.save(output_file)
                wb.close()

            time.sleep(interval_seconds)
    except KeyboardInterrupt:
        print("Monitor detenido por usuario.")
    except Exception as e:
        print(f"Monitor finalizado por error: {e}")

if __name__ == "__main__":
    monitor_match_and_log_excel(interval_seconds=5, output_file="registro.xlsx")
