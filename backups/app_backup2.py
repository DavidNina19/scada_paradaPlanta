from redis_connect import MQTTWithRedis
from api_alex_connect import Api
import pandas as pd
import time
from datetime import datetime
import re
import os
from openpyxl import Workbook, load_workbook
#from datetime import timedelta


def datos_alex_proceso():
    url = 'http://192.168.252.6/serviciowebaccesonet/api/authentication/logintercero'
    filas = []
    for i in range(1, 3):
        urlMaquina = f'http://192.168.252.6/serviciowebtercerosnet/api/MaquinaOrdenParte/{i}'
        api = Api(url, urlMaquina)
        data = api.get_all_data()

        if not data:
            continue

        import pandas as pd
        # Normalizar distintos tipos de retorno
        if isinstance(data, pd.DataFrame):
            filas.extend(data.to_dict('records'))
        elif isinstance(data, dict):
            filas.append(data)
        else:  # lista de dicts u otros iterables
            try:
                filas.extend(list(data))
            except Exception:
                filas.append(data)

    if not filas:
        return None

    import pandas as pd
    df = pd.DataFrame(filas)
    return df

def datos_paradas_scada():
    paradaScada = MQTTWithRedis()
    df = paradaScada.paradas_produccion()
    return df

def extraer_datos_alex():
    df = datos_alex_proceso()
    columnas_objetivo = ['codMaquina','idNumOrd', 'inicioParada', 'desParada', 'operario']
    NOMBRE_COLUMNA_FECHA_FINAL = 'inicioParada_alex'

    if df is None or df.empty:
        cols_finales = [c if c != 'inicioParada' else NOMBRE_COLUMNA_FECHA_FINAL for c in columnas_objetivo]
        return pd.DataFrame(columns=cols_finales)

    # --- Lógica de mapeo de columnas omitida por brevedad, asume que mapea a 'columnas_objetivo' ---
    # ... (Tu lógica de mapeo flexible aquí) ...
    cols_map = {}
    cols_lower = {c.lower(): c for c in df.columns}
    for target in columnas_objetivo:
        t_lower = target.lower()
        if t_lower in cols_lower:
            cols_map[target] = cols_lower[t_lower]
            continue
        candidates = [c for c in df.columns if t_lower in c.lower()]
        cols_map[target] = candidates[0] if candidates else None

    result = {}
    for target, src in cols_map.items():
        if src:
            result[target] = df[src]
        else:
            result[target] = pd.Series([pd.NA] * len(df), index=df.index)

    df_res = pd.DataFrame(result).reset_index(drop=True)

    if 'inicioParada' in df_res.columns:
        df_res['inicioParada'] = pd.to_datetime(df_res['inicioParada'], errors='coerce')
    
    if 'inicioParada' in df_res.columns:
        df_res = df_res.rename(columns={'inicioParada': NOMBRE_COLUMNA_FECHA_FINAL})

    return df_res

def match_alex_scada_paradas():
    df_alex = extraer_datos_alex()
    df_scada = datos_paradas_scada()

    # columnas objetivo de salida
    out_cols = [
        'codMaquina', 'inicioParada_alex', 'desParada', 'operario', 'idNumOrd',
        'inicioParada', 'finParada', 'horaParada', 'fecha', 'horaInicio', 'horaFin', 'turno'
    ]

    if df_alex is None or df_alex.empty or df_scada is None or df_scada.empty:
        return pd.DataFrame(columns=out_cols)
    
    df_matches = pd.merge(
    df_alex, 
    df_scada, 
    on='codMaquina', # La columna clave para la comparación
    how='inner'      # Tipo de unión: 'inner' para encontrar la intersección (coincidencias)
    )
    return df_matches[out_cols]
    

def monitor_match_and_log_excel(interval_seconds: int = 5, output_file: str = "registro_app2.xlsx", sheet_name: str = "registro_app2"):
    seen = set()
    try:
        while True:
            df_matches = match_alex_scada_paradas()
            if df_matches is None:
                df_matches = pd.DataFrame()

            rows = [tuple(r) for r in df_matches.itertuples(index=False, name=None)]
            new_rows = [r for r in rows if r not in seen]

            if new_rows:
                # marcar filas nuevas como vistas
                for r in new_rows:
                    seen.add(r)

                # preparar libro/hoja
                if not os.path.exists(output_file):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name
                    write_header = True
                else:
                    wb = load_workbook(output_file)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)
                    # si la hoja está vacía (sin filas), escribimos encabezado
                    write_header = (ws.max_row == 0)

                # Escribir header solo si se creó el archivo/hoja o está vacío
                if write_header and not df_matches.empty:
                    ws.append(list(df_matches.columns))

                # Escribir únicamente las filas nuevas (new_rows)
                for r in new_rows:
                    row_vals = []
                    for v in r:
                        try:
                            if pd.isna(v):
                                row_vals.append("")
                            elif hasattr(v, "strftime"):
                                try:
                                    row_vals.append(v.strftime("%Y-%m-%d %H:%M:%S"))
                                except Exception:
                                    row_vals.append("")
                            else:
                                row_vals.append(v)
                        except Exception:
                            try:
                                row_vals.append(str(v))
                            except Exception:
                                row_vals.append("")
                    ws.append(row_vals)

                wb.save(output_file)
                wb.close()

            time.sleep(interval_seconds)
    except KeyboardInterrupt:
        print("Monitor detenido por usuario.")
    except Exception as e:
        print(f"Monitor finalizado por error: {e}")

if __name__ == "__main__":
    df_matches = match_alex_scada_paradas()
    print(df_matches)
    monitor_match_and_log_excel(interval_seconds=5, output_file="registro_app2.xlsx")
    #df_alex = extraer_datos_alex()
    #df_scada = datos_paradas_scada()
    #print("Datos Alex:")
    #print(df_alex)
    #print("\nDatos SCADA:")
    #print(df_scada)
