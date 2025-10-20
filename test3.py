from redis_connect import MQTTWithRedis
from api_alex_connect import Api
import pandas as pd
import time
from datetime import datetime
import re
import os
from openpyxl import Workbook, load_workbook
#from datetime import timedelta


def datos_alex_proceso():
    url = 'http://192.168.252.6/serviciowebaccesonet/api/authentication/logintercero'
    filas = []
    for i in range(1, 3):
        urlMaquina = f'http://192.168.252.6/serviciowebtercerosnet/api/MaquinaOrdenParte/{i}'
        api = Api(url, urlMaquina)
        data = api.get_all_data()

        if not data:
            continue

        import pandas as pd
        # Normalizar distintos tipos de retorno
        if isinstance(data, pd.DataFrame):
            filas.extend(data.to_dict('records'))
        elif isinstance(data, dict):
            filas.append(data)
        else:  # lista de dicts u otros iterables
            try:
                filas.extend(list(data))
            except Exception:
                filas.append(data)

    if not filas:
        return None

    import pandas as pd
    df = pd.DataFrame(filas)
    return df

def datos_paradas_scada():
    paradaScada = MQTTWithRedis()
    df = paradaScada.paradas_produccion()
    return df

def extraer_datos_alex():
    df = datos_alex_proceso()
    columnas_objetivo = ['codMaquina','idNumOrd', 'inicioParada', 'desParada', 'operario']
    NOMBRE_COLUMNA_FECHA_FINAL = 'inicioParada_alex'

    if df is None or df.empty:
        cols_finales = [c if c != 'inicioParada' else NOMBRE_COLUMNA_FECHA_FINAL for c in columnas_objetivo]
        return pd.DataFrame(columns=cols_finales)

    # --- Lógica de mapeo de columnas omitida por brevedad, asume que mapea a 'columnas_objetivo' ---
    # ... (Tu lógica de mapeo flexible aquí) ...
    cols_map = {}
    cols_lower = {c.lower(): c for c in df.columns}
    for target in columnas_objetivo:
        t_lower = target.lower()
        if t_lower in cols_lower:
            cols_map[target] = cols_lower[t_lower]
            continue
        candidates = [c for c in df.columns if t_lower in c.lower()]
        cols_map[target] = candidates[0] if candidates else None

    result = {}
    for target, src in cols_map.items():
        if src:
            result[target] = df[src]
        else:
            result[target] = pd.Series([pd.NA] * len(df), index=df.index)

    df_res = pd.DataFrame(result).reset_index(drop=True)

    if 'inicioParada' in df_res.columns:
        df_res['inicioParada'] = pd.to_datetime(df_res['inicioParada'], errors='coerce')
    
    if 'inicioParada' in df_res.columns:
        df_res = df_res.rename(columns={'inicioParada': NOMBRE_COLUMNA_FECHA_FINAL})

    return df_res

def match_alex_scada_paradas():
    df_alex = extraer_datos_alex()
    df_scada = datos_paradas_scada()

    # columnas objetivo de salida
    out_cols = [
        'codMaquina', 'inicioParada_alex', 'desParada', 'operario', 'idNumOrd',
        'inicioParada', 'finParada', 'horaParada', 'fecha', 'horaInicio', 'horaFin', 'turno'
    ]

    if df_alex is None or df_alex.empty or df_scada is None or df_scada.empty:
        return pd.DataFrame(columns=out_cols)

    # merge para tener datos combinados por codMaquina
    df_matches = pd.merge(
        df_alex,
        df_scada,
        on='codMaquina',
        how='inner'
    )

    # Asegurar que las columnas de tiempo estén en datetime
    if 'inicioParada_alex' in df_matches.columns:
        df_matches['inicioParada_alex'] = pd.to_datetime(df_matches['inicioParada_alex'], errors='coerce')
    if 'inicioParada' in df_matches.columns:
        df_matches['inicioParada'] = pd.to_datetime(df_matches['inicioParada'], errors='coerce')
    if 'finParada' in df_matches.columns:
        df_matches['finParada'] = pd.to_datetime(df_matches['finParada'], errors='coerce')

    # df_filtered: registros donde inicioParada_alex está dentro del rango [inicioParada, finParada]
    mask = pd.Series(False, index=df_matches.index)
    if {'inicioParada_alex', 'inicioParada', 'finParada'}.issubset(df_matches.columns):
        a = df_matches['inicioParada_alex']
        s = df_matches['inicioParada']
        f = df_matches['finParada']
        mask = (pd.notna(a)) & (
            ((pd.notna(s) & pd.notna(f) & (a >= s) & (a <= f))) |
            ((pd.notna(s) & pd.isna(f) & (a >= s)))
        )

    df_filtered = df_matches[mask].copy()

    # df_copy: copia del df_scada con las columnas out_cols; inicioParada_alex y desParada vacíos
    df_copy = df_scada.copy()
    # si df_copy usa 'codmaq' renombrar para mantener 'codMaquina'
    if 'codMaquina' not in df_copy.columns and 'codmaq' in df_copy.columns:
        df_copy = df_copy.rename(columns={'codmaq': 'codMaquina'})

    # asegurar las columnas out_cols existen en df_copy
    for col in out_cols:
        if col not in df_copy.columns:
            df_copy[col] = pd.NA

    # inicializar campos vacíos para rellenar
    df_copy['inicioParada_alex'] = pd.NaT
    df_copy['desParada'] = pd.NA

    # convertir columnas de tiempo en df_copy para comparaciones
    if 'inicioParada' in df_copy.columns:
        df_copy['inicioParada'] = pd.to_datetime(df_copy['inicioParada'], errors='coerce')
    else:
        df_copy['inicioParada'] = pd.NaT
    if 'finParada' in df_copy.columns:
        df_copy['finParada'] = pd.to_datetime(df_copy['finParada'], errors='coerce')
    else:
        df_copy['finParada'] = pd.NaT

    # recorrer df_filtered e insertar/enriquecer en df_copy según la condición por máquina y rango
    for _, row in df_filtered.iterrows():
        cod = row.get('codMaquina')
        a_inicio = row.get('inicioParada_alex')
        a_des = row.get('desParada', pd.NA)
        a_id = row.get('idNumOrd', row.get('idNumOrden', pd.NA))
        a_oper = row.get('operario', pd.NA)

        if pd.isna(a_inicio) or pd.isna(cod):
            continue

        same_machine = df_copy['codMaquina'] == cod

        in_range = same_machine & (
            ((pd.notna(df_copy['inicioParada']) & pd.notna(df_copy['finParada']) &
              (df_copy['inicioParada'] <= a_inicio) & (a_inicio <= df_copy['finParada']))) |
            ((pd.notna(df_copy['inicioParada']) & df_copy['finParada'].isna() & (a_inicio >= df_copy['inicioParada'])))
        )

        # fallback: si no hay en_range, intentar emparejar por máquina y fecha/hora aproximada
        if not in_range.any():
            in_range = same_machine & pd.notna(df_copy['inicioParada']) & (a_inicio >= df_copy['inicioParada'])

        # asignar valores en las filas encontradas
        df_copy.loc[in_range, 'inicioParada_alex'] = a_inicio
        df_copy.loc[in_range, 'desParada'] = a_des
        df_copy.loc[in_range, 'idNumOrd'] = a_id
        df_copy.loc[in_range, 'operario'] = a_oper

    # devolver solo las columnas de salida (existentes)
    cols_to_return = [c for c in out_cols if c in df_copy.columns]
    return df_copy[cols_to_return].reset_index(drop=True)
    

def monitor_match_and_log_excel(interval_seconds: int = 5, output_file: str = "test3.xlsx", sheet_name: str = "test3"):
    seen = set()
    try:
        while True:
            df_matches = match_alex_scada_paradas()
            if df_matches is None:
                df_matches = pd.DataFrame()

            rows = [tuple(r) for r in df_matches.itertuples(index=False, name=None)]
            new_rows = [r for r in rows if r not in seen]

            if new_rows:
                # marcar filas nuevas como vistas
                for r in new_rows:
                    seen.add(r)

                # preparar libro/hoja
                if not os.path.exists(output_file):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name
                    write_header = True
                else:
                    wb = load_workbook(output_file)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)
                    # si la hoja está vacía (sin filas), escribimos encabezado
                    write_header = (ws.max_row == 0)

                # Escribir header solo si se creó el archivo/hoja o está vacío
                if write_header and not df_matches.empty:
                    ws.append(list(df_matches.columns))

                # Escribir únicamente las filas nuevas (new_rows)
                for r in new_rows:
                    row_vals = []
                    for v in r:
                        try:
                            if pd.isna(v):
                                row_vals.append("")
                            elif hasattr(v, "strftime"):
                                try:
                                    row_vals.append(v.strftime("%Y-%m-%d %H:%M:%S"))
                                except Exception:
                                    row_vals.append("")
                            else:
                                row_vals.append(v)
                        except Exception:
                            try:
                                row_vals.append(str(v))
                            except Exception:
                                row_vals.append("")
                    ws.append(row_vals)

                wb.save(output_file)
                wb.close()

            time.sleep(interval_seconds)
    except KeyboardInterrupt:
        print("Monitor detenido por usuario.")
    except Exception as e:
        print(f"Monitor finalizado por error: {e}")

if __name__ == "__main__":
    df_matches = match_alex_scada_paradas()
    print(df_matches)
    monitor_match_and_log_excel(interval_seconds=5, output_file="test3.xlsx")
    #df_alex = extraer_datos_alex()
    #df_scada = datos_paradas_scada()
    #print("Datos Alex:")
    #print(df_alex)
    #print("\nDatos SCADA:")
    #print(df_scada)
